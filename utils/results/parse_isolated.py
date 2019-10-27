from __future__ import division

import openpyxl
import numpy as np
import matplotlib.pyplot as plt


def gain(value, base):
    return 100 * (value - base) / base


if __name__ == '__main__':
    wb = openpyxl.load_workbook('/home/srishtid/Dropbox/Adaptive-Traffic-Signal-Control/ICAPS-2019/Results/FinalResults.xlsx')

    isolated = wb[wb.sheetnames[4]]
    print(isolated)

    samples = [1, 5, 10, 20, 30]
    demands = ['900V', '1200V', '1500V']
    start_row = [6, 36, 66]
    saa_start_col = 5

    delays_over_all_demands = []
    
    for index, demand in enumerate(demands):
        all_delays = []

        usurtrac = []
        for case in range(20):
            delay = isolated.cell(row=start_row[index]+case, column=2).value
            # print(delay)
            if type(delay) != float:
                continue
            usurtrac.append(delay)
        all_delays.append(usurtrac)

        csurtrac = []
        for case in range(20):
            delay = isolated.cell(row=start_row[index]+case, column=3).value
            if type(delay) != float:
                continue
            csurtrac.append(delay)
        all_delays.append(csurtrac)

        for sample_index, sample in enumerate(samples):
            for config_index, config in enumerate(['uncoo', 'coo']):                     
                delays = []
                for set_index in range(5):
                    col = saa_start_col + 25 * sample_index + 5 * set_index + 2 * config_index
                    for row in range(start_row[index], start_row[index]+20):
                        delay = isolated.cell(row=row,column=col).value
                        if type(delay) != float:
                            continue
                        delays.append(delay)

                print('Delays for S{} {}'.format(sample, config))
                print(delays)
                all_delays.append(delays)

        delays_over_all_demands.append(all_delays)

    # print delays_over_all_demands[0]

    figure, axes = plt.subplots(1, 3, sharex='all', sharey='all', figsize=(15, 5))
    for index, demand in enumerate(demands):
        delays = delays_over_all_demands[index]
        axis = axes[index]
        axis.boxplot([delay for delay in delays], widths=0.2)
        axis.set_title(demand)
        axis.set_ylabel('Delay (seconds)')
        plt.setp(axis.get_xticklabels(), rotation=90)
       
    plt.setp(axes, xticks=range(1,13), xticklabels=['u-surtrac', 'c-surtrac',
                                                   'u-S1', 'c-S1',
                                                   'u-S5', 'c-S5',
                                                   'u-S10', 'c-S10',
                                                   'u-S20', 'c-S20',
                                                   'u-S30', 'c-S30'])
    
    plt.show()

    # gains = []
    # for case in range(20):
    #     gains.append(gain(value=delays[case], base=surtrac[case]))
    # print('Average gain for S{}-R{} = {}'.format(sample, set_index+1, np.mean(gains)))

    # print(surtrac)
    # print(saa)
